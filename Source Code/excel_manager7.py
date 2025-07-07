# File: excel_manager7.py
# Purpose: Manage the Excel file for literature organization, including loading, appending data,
# and formatting the Excel sheet with hyperlinks and proper styles.

# Copyright 2025 Aarush Jhaveri
# Copyright 2025 Goutam Narayan Tumulu
# Copyright 2025 Sanjay Mahajani
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import config
from openpyxl.styles import Alignment
from datetime import datetime
from citation_parser2 import parse_citation_file
from openpyxl.styles import Font

def load_or_create_excel(file_path):
    """Load or create an Excel file, ensuring proper headers."""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        existing_df = pd.read_excel(file_path)
        existing_titles = set(existing_df['Title'].dropna().unique()) if 'Title' in existing_df.columns else set()
    except Exception:
        print("⚠️ Error loading Excel file. Creating a new one...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Literature Organisation"
        ws.append(config.EXCEL_HEADERS)
        wb.save(file_path)
        existing_titles = set()
    
    return wb, ws

def append_data_to_excel(file_path, ws, wb, new_data, target_directory, citation_paths):
    """Append new citation data while preserving formatting."""
    new_df = pd.DataFrame(new_data)[config.EXCEL_HEADERS]

    if 'Sr No.' not in new_df.columns:
        new_df['Sr No.'] = None

    # Convert 'Access Date' to datetime for sorting
    new_df['Access Date'] = pd.to_datetime(new_df['Access Date'], format='%b %d %Y')
    
    # Format 'Access Date' to the desired format
    new_df['Access Date'] = new_df['Access Date'].dt.strftime('%b %d %Y')
    
    # Define column order based on headers
    column_order = config.EXCEL_HEADERS

    new_df['Sr. No.'] = range(1, len(new_df) + 1)
    
    for _, row in new_df.iterrows():
        row_list = row.tolist()
        
        # Get Access Date and create the hyperlink path
        access_date = row['Access Date']
        date_folder_path = os.path.join(target_directory, access_date)
        
        # Add hyperlink for Access Date
        row_list[column_order.index('Access Date')] = f'=HYPERLINK("{date_folder_path}", "{access_date}")'

        # Add hyperlink for Title
        title = row['Title']
        citation_filename = citation_paths.get(title, "")

        if citation_filename:
            # Build the full citation path 
            access_date = row['Access Date']
            date_folder_path = os.path.join(target_directory, access_date)
            full_citation_path = os.path.join(date_folder_path, citation_filename)

            # Add the hyperlink to the Title
            row_list[config.EXCEL_HEADERS.index('Title')] = f'=HYPERLINK("{full_citation_path}", "{title}")'
        
        ws.append(row_list)
        
        sr_no_cell = ws.cell(row=ws.max_row, column=column_order.index('Sr. No.') + 1)
        sr_no_cell.value = ws.max_row - 1

    # Add hyperlinks to DOI columns
    if 'DOI' in new_df.columns:
        new_df['DOI'] = new_df['DOI'].apply(lambda x: f'=HYPERLINK("{x}", "{x}")' if isinstance(x, str) and x.startswith("http") else x)
    
    # Preserve column width and enable text wrapping for the Abstract and Title columns
    for col_idx, column_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        column_name = ws.cell(row=1, column=col_idx).value
        
        if column_name == "Abstract":
            ws.column_dimensions[col_letter].width = 75  # Set Abstract column width to 75
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping
        elif column_name == "Title":
            ws.column_dimensions[col_letter].width = 36  # Set Title column width to 36  
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping
        elif column_name == "Access Date":
            ws.column_dimensions[col_letter].width = 20  # Set Access Date column width to 20      
        else:
            # Auto-adjust width for other columns
            max_length = 0
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length  # Adjust column width dynamically
    
    # Apply hyperlink styling after appending rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Skip header row
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK('):
                cell.font = Font(color="0563C1", underline="single")  # Excel hyperlink color + underline
    
    wb.save(file_path)
    print(f"✅ Appended {len(new_df)} new entries to {file_path}")
